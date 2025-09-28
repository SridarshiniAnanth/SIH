
import os
import time
import json
import datetime
import base64
import threading
import cv2
import numpy as np
from io import BytesIO
from PIL import Image
from flask import Flask, request, render_template, jsonify, send_file
import openpyxl
import folium
import smtplib
from email.message import EmailMessage



def send_mail(subject, body, to_email, from_email, from_password, smtp_server="smtp.gmail.com", smtp_port=587):
    
    if isinstance(to_email, str):
        to_email = [to_email]

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = ", ".join(to_email)
    msg.set_content(body)  
    
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls() 
            server.login(from_email, from_password)
            server.send_message(msg)
            print("✅ Email sent successfully to", to_email)
            return True
    except Exception as e:
        print("❌ Error sending email:", e)
        return False

DB_DIR = "db"
MODEL_PATH = "face_model.yml"
LABELS_PATH = "labels.json"

os.makedirs(DB_DIR, exist_ok=True)


face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
recognizer = cv2.face.LBPHFaceRecognizer_create()


SAMPLES_PER_USER = 10
MIN_FACE_SIZE = 100
SHARPNESS_THRESHOLD = 60.0
RESIZE_TO = (200, 200)
CONFIDENCE_THRESHOLD = 60.0
REQUIRED_MATCHES = 3
VERIFY_TIMEOUT = 12.0


def variance_of_laplacian(image_gray):
    return cv2.Laplacian(image_gray, cv2.CV_64F).var()

def save_labels(label_map):
    with open(LABELS_PATH, "w", encoding="utf-8") as f:
        json.dump(label_map, f, ensure_ascii=False, indent=2)

def load_labels():
    if os.path.exists(LABELS_PATH):
        with open(LABELS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def decode_base64_to_gray(data_url):
    try:
        _, encoded = data_url.split(",", 1)
        img_bytes = base64.b64decode(encoded)
        img = Image.open(BytesIO(img_bytes)).convert("L")
        return np.array(img)
    except Exception:
        return None


STUDENT_LIST_XLSX = "student_list.xlsx"  

def get_usermail(username):
    """Fetch usermail from student_list.xlsx by username"""
    if not os.path.exists(STUDENT_LIST_XLSX):
        return None
    wb = openpyxl.load_workbook(STUDENT_LIST_XLSX)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        uname, mail, _ = row
        if uname == username:
            return mail
    return None

def ensure_user_dirs(username,usermail):
    """Create dataset/login folders and login.xlsx for a user, and update student_list.xlsx"""
    user_dir = os.path.join(DB_DIR, username)
    dataset_dir = os.path.join(user_dir, "dataset")
    login_dir = os.path.join(user_dir, "login")
    xlsx_path = os.path.join(user_dir, "login.xlsx")

    os.makedirs(dataset_dir, exist_ok=True)
    os.makedirs(login_dir, exist_ok=True)

   
    if not os.path.exists(xlsx_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["timestamp", "photo_path", "location_map_path", "avg_confidence", "latitude", "longitude"])
        wb.save(xlsx_path)


    if not os.path.exists(STUDENT_LIST_XLSX):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["username","mail" , "create_date"])
        wb.save(STUDENT_LIST_XLSX)

    
    wb = openpyxl.load_workbook(STUDENT_LIST_XLSX)
    ws = wb.active
    usernames = [row[0].value for row in ws.iter_rows(min_row=2)]
    if username not in usernames:
        ws.append([username,usermail, datetime.datetime.now().isoformat()])
        wb.save(STUDENT_LIST_XLSX)

    return dataset_dir, login_dir, xlsx_path


def train_model():
    faces, labels = [], []
    label_map = {}
    current_id = 0
    for username in sorted(os.listdir(DB_DIR)):
        dataset_dir = os.path.join(DB_DIR, username, "dataset")
        if not os.path.isdir(dataset_dir):
            continue
        for file in os.listdir(dataset_dir):
            if not file.lower().endswith(".jpg"):
                continue
            path = os.path.join(dataset_dir, file)
            gray = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
            if gray is None:
                continue
            if username not in label_map:
                label_map[username] = current_id
                current_id += 1
            if gray.shape != RESIZE_TO:
                try:
                    gray = cv2.resize(gray, RESIZE_TO)
                except Exception:
                    continue
            faces.append(gray)
            labels.append(label_map[username])
    if faces and labels:
        recognizer.train(faces, np.array(labels))
        recognizer.write(MODEL_PATH)
        save_labels(label_map)
        print("✅ Model trained and saved.")
    else:
        print("⚠️ No faces found to train.")
    return label_map


reg_lock = threading.Lock()
ver_lock = threading.Lock()
reg_progress = {}
ver_progress = {}
latest_locations = {}  

def cleanup_progress(interval=300):
    while True:
        now = time.time()
        with reg_lock:
            for k in list(reg_progress.keys()):
                if now - reg_progress[k].get("last_seen", 0) > interval:
                    del reg_progress[k]
        with ver_lock:
            for k in list(ver_progress.keys()):
                if now - ver_progress[k].get("last_seen", 0) > interval:
                    del ver_progress[k]
        time.sleep(interval)

threading.Thread(target=cleanup_progress, daemon=True).start()


app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/register_frame", methods=["POST"])
def register_frame():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    usermail = (data.get("usermail") or "").strip()
    frame_b64 = data.get("frame")
    if not username or not frame_b64:
        return jsonify({"status":"error","message":"username and frame required"}), 400
    dataset_dir, _, _ = ensure_user_dirs(username,usermail)
    gray = decode_base64_to_gray(frame_b64)
    if gray is None:
        return jsonify({"status":"error","message":"invalid image"}), 400
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5)
    if len(faces) == 0:
        with reg_lock:
            ent = reg_progress.setdefault(username, {"count":0})
            ent["last_seen"] = time.time()
        return jsonify({"status":"ok","message":"no face","saved_count": ent["count"], "required": SAMPLES_PER_USER})
    x,y,w,h = sorted(faces, key=lambda r: r[2]*r[3], reverse=True)[0]
    if w < MIN_FACE_SIZE or h < MIN_FACE_SIZE:
        return jsonify({"status":"ok","message":"face too small","saved_count": reg_progress.get(username,{}).get("count",0), "required": SAMPLES_PER_USER})
    face_img = gray[y:y+h, x:x+w]
    if variance_of_laplacian(face_img) < SHARPNESS_THRESHOLD:
        return jsonify({"status":"ok","message":"blurry","saved_count": reg_progress.get(username,{}).get("count",0), "required": SAMPLES_PER_USER})
    face_resized = cv2.resize(face_img, RESIZE_TO)
    face_resized = cv2.equalizeHist(face_resized)
    with reg_lock:
        ent = reg_progress.setdefault(username, {"count":0, "last_seen": time.time()})
        ent["count"] += 1
        ent["last_seen"] = time.time()
        saved_count = ent["count"]
    fname = f"{saved_count:02d}.jpg"
    cv2.imwrite(os.path.join(dataset_dir, fname), face_resized)
    if saved_count >= SAMPLES_PER_USER:
        label_map = train_model()
        with reg_lock:
            reg_progress.pop(username, None)
        return jsonify({"status":"done","message":"registration complete","saved_count": saved_count, "labels": label_map})
    return jsonify({"status":"ok","message":"saved","saved_count": saved_count, "required": SAMPLES_PER_USER})


@app.route("/verify_frame", methods=["POST"])
def verify_frame():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    frame_b64 = data.get("frame")
    latitude = data.get("latitude")
    longitude = data.get("longitude")

    if not username or not frame_b64:
        return jsonify({"status": "error", "message": "username and frame required"}), 400
    if not os.path.exists(MODEL_PATH):
        return jsonify({"status": "fail", "message": "No trained model. Register first."})

    dataset_dir, login_dir, xlsx_path = ensure_user_dirs(username, "")
    gray = decode_base64_to_gray(frame_b64)
    if gray is None:
        return jsonify({"status": "error", "message": "invalid image"}), 400
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5)
    if len(faces) == 0:
        return jsonify({"status": "ok", "message": "no face", "matches": 0})

    x, y, w, h = sorted(faces, key=lambda r: r[2] * r[3], reverse=True)[0]
    if w < MIN_FACE_SIZE or h < MIN_FACE_SIZE:
        return jsonify({"status": "ok", "message": "face too small", "matches": 0})

    face_img = cv2.resize(gray[y:y + h, x:x + w], RESIZE_TO)
    face_img = cv2.equalizeHist(face_img)

    label_map = load_labels()
    if username not in label_map:
        return jsonify({"status": "fail", "message": "Username not registered"})

    recognizer.read(MODEL_PATH)
    label, confidence = recognizer.predict(face_img)
    expected_label = int(label_map[username])

    with ver_lock:
        ent = ver_progress.setdefault(username, {"matches": 0, "confidences": [], "start_time": time.time(), "last_seen": time.time()})
        ent["last_seen"] = time.time()

        if label == expected_label and confidence < CONFIDENCE_THRESHOLD:
            ent["matches"] += 1
            ent["confidences"].append(confidence)
        else:
            ent["matches"] = 0
            ent["confidences"] = []

       
        if ent["matches"] >= REQUIRED_MATCHES:
            avg_conf = float(np.mean(ent["confidences"]))
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            folder = os.path.join(login_dir, f"unlocked_{ts}")
            os.makedirs(folder, exist_ok=True)

            photo_path = os.path.join(folder, "photo.jpg")
            cv2.imwrite(photo_path, gray)

          
            map_img_path = os.path.join(folder, "location_map.png")
            if latitude is not None and longitude is not None:
                m = folium.Map(location=[latitude, longitude], zoom_start=16)
                folium.Marker([latitude, longitude], tooltip=username).add_to(m)
                m.save(os.path.join(folder, "map.html"))

            location_txt_path = os.path.join(folder, "location.txt")
            with open(location_txt_path, "w") as f:
                f.write(f"Latitude:{latitude}\nLongitude:{longitude}\nLogin-Time:{datetime.datetime.now().strftime('%d/%m/%y - %H:%M:%S : %f')[:-3] }\nAccuracy:{avg_conf}")

           
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            ws.append([datetime.datetime.now().isoformat(), photo_path, map_img_path, avg_conf, latitude, longitude])
            wb.save(xlsx_path)
#                                                          
         
            send_mail(
                subject="Attendance Verified",
                body=f"Hello {username}, your login was verified your Attendance successfully ✅\n Time: {datetime.datetime.now().isoformat()}\nConfidence: {avg_conf:.2f}\nLatitude: {latitude}, Longitude: {longitude}",
                to_email=get_usermail(username),
                from_email="sujiths812006@gmail.com",
                from_password=""
            )

            ver_progress.pop(username, None)
            return jsonify({
                "status": "verified",
                "message": f"✅ Verified for {username} (avg_conf={avg_conf:.2f})",
                "photo_path": photo_path,
                "map_path": map_img_path,
                "latitude": latitude,
                "longitude": longitude
            })

        if time.time() - ent["start_time"] > VERIFY_TIMEOUT:
            ver_progress.pop(username, None)
            return jsonify({"status": "failed_final", "message": "❌ Verification failed (timeout)"})

   
        return jsonify({
            "status": "ok",
            "message": f"Partial matches {ent['matches']}/{REQUIRED_MATCHES}",
            "matches": ent["matches"]
        })


@app.route("/location", methods=["POST"])
def location():
    data = request.get_json()
    username = data.get("username")
    if not username:
        return {"status":"fail", "message":"username required"}, 400
    latest_locations[username] = {"latitude": data.get("latitude"), "longitude": data.get("longitude")}
    return {"status":"ok", "received": data}


@app.route("/train", methods=["POST"])
def route_train():
    label_map = train_model()
    return jsonify({"status":"ok","message":"retrained","labels":label_map})

if __name__ == "__main__":
    context = None
    if os.path.exists("cert.pem") and os.path.exists("key.pem"):
        context = ("cert.pem", "key.pem")
    app.run(host="0.0.0.0", port=5000, ssl_context=context)
